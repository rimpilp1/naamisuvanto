
from django.shortcuts import render
from django.http import HttpResponse
from django.template import loader
from .forms import SaalisForm
from .models import Saalis
import datetime
from django.utils import timezone


from PIL import Image
import requests
from io import BytesIO
from django.conf import settings
from django.core.paginator import Paginator

from django.shortcuts import redirect
from django.db import transaction
from django.utils.datastructures import MultiValueDictKeyError
from django.core.files import File
from saalistilastot.models import Saalis
from openpyxl import load_workbook
from openpyxl import Workbook
import uuid
import sys, traceback 
import openpyxl
import urllib
import os
import traceback 


def delete(request):
    #Don't use if you don't know what this does!
    #Should only be used for quick way to delete the whole database.
    return HttpResponse("This feature is disabled")
    Saalis.objects.all().delete()
    return HttpResponse("All have been deleted")

def parse(request):
        #DO NOT USE THIS FUNCTION WITHOUT KNOWING WHAT IT DOES. 
        #

        
        return HttpResponse("This feature is disabled")
        
        #NOT NULL AIKA, PAINO, PAIKKA, SAAJA
        #Store the cases where continue is used as their own worksheet. 
        #Using this function might require manual adjustment of the code.
        
        
        PROJECT_ROOT = os.path.abspath(os.path.dirname(__file__))
        excel_file = open(os.path.join(PROJECT_ROOT, 'parse.xlsx'))
        print(os.path.join(PROJECT_ROOT, 'parse.xlsx'))
        
        wb = load_workbook((os.path.join(PROJECT_ROOT, 'parse.xlsx')))
        print(wb.sheetnames)

        wb2 = Workbook()
        ws2 = wb2.active
        #data = xlsx_get(excel_file)
        saaliit = wb.active

        wrong_kalat = set()
        sukupuolet = set()
        skip_paino = []
        skip_pituus = []
        skip_image = []
        null = {}
        skip = False
        skip_count = 0
        count = 0
        print("________")

        for row in saaliit.values:
            skip = False
            fault = []
            print(count)
            count += 1
            #print("------")
            if row[0] == 'Pvm':
                skip = True
                #SKIP -> STORE results in new file

            #print(row[0],type(row[0]))
            #print(row[1])
            #print(row[2])
            #print(row[3])
            #print(row[4])
            #print(row[5])
            #print(row[6])
            #print(row[7])
            #print(row[8])
            
            objects = []
            aika = row[0]
            kuva = row[1]
            kala = row[2]
            sp = row[3]
            paino = row[4]
            pituus = row[5]
            paikka = row[6]
            viehe = row[7]
            kalastaja = row[8]

            for i in range(0,9):
                if row[i] == None:
                    if i in null:
                        null[i] = null[i]+1
                    else:
                        null[i] = 1
            
            #Normal processing
            
            """
            if kala == None:
                wrong_kalat.add(kala)
                fault.append(2)
                skip = True
            elif ('lohi' not in kala.lower()):
                wrong_kalat.add(kala.lower())
                fault.append(2)
                skip = True
                #SKIP -> STORE results in new file
                """

            #Faulty processing
            
            if kala == None:
                kala = 'lohi'
            elif 'lohi' in kala.lower():
                kala = 'lohi'
            elif 'taimen' in kala.lower():
                kala = 'taimen'
            elif 'laji' in kala.lower():
                skip = True
                wrong_kalat.add(kala.lower())
            else:
                print(kala)
                return HttpResponse("kala problem")

            if 'lohi' in kala.lower() and 'taimen' in kala.lower():
                print(kala)
                return HttpResponse("kala problem")
            

            vapaus = False

            if kala == None:
                True
            elif ('cr' in kala.lower()):
                vapaus = True
            elif 'c' in kala.lower():
                return HttpResponse("catch and release problem")


            if (row[3] == None):
                sp = "-"
            else:
                sp = row[3].strip().upper()

            if (sp != "N" and sp != "U"):
                sukupuolet.add(sp)

            

            if (kuva == "" or kuva == None or 'http' not in kuva):
                kuva = None
            else:
                #print(row[1])
                #print(kuva)
                #print(settings.MEDIA_ROOT)
                #print("----")
                try:
                
                    kuva = kuva[kuva.find('http'):]
                    #response = requests.get(kuva)
                    #kuva = Image.open(BytesIO(response.content))
                    root = os.path.join(settings.MEDIA_ROOT, "images")
                    
                    name = kuva[(kuva.rfind('/')+1):]
                    c_i = name.rfind('.')
                    name = name[:c_i] + str(uuid.uuid4()) + name[c_i:] 
                    root = os.path.join(root, name)
                    kuva = urllib.request.urlretrieve(kuva,root)
                
                
                except:
                   #print("fail")
                    fault.append(1)
                    skip_image.append(kuva)
                    skip = True
                    
            

            if paino == None:
                True
                #Fault disable, Normal disable 
                #fault.append(4)
                #skip = True
            elif type(paino) != type(10) and type(paino) != type(0.1):
                skip_paino.append(paino)
                fault.append(4)
                skip = True
            elif paino >= 30:
                fault.append(4)
                skip_paino.append(paino)
                skip = True


            if pituus == None:
                True
            elif type(pituus) != type(10) and type(pituus) != type(0.1):
                skip_pituus.append(pituus)
                fault.append(5)
                skip = True

            # normal 50, fautly 40 and Note None
            #elif pituus < 50:
            elif pituus < 40 and pituus != None:
                skip_pituus.append(pituus)
                fault.append(5)
                skip = True

            if paikka == None:
                fault.append(6)
                skip = True

            if viehe == None or viehe == "":
                viehe = '-'

            if kalastaja == None:
                fault.append(8)
                skip = True

            test = False
            
            if skip:
                skip_count += 1
                #str(fault)
                tmp = row + tuple(fault)
                #print("------")
                #print(tmp)
                #print(row)
                ws2.append(tmp)

            else:
                #try:
                    #print(str(type(kuva)))
                if kuva != None:
                    test = True
                    kuva = kuva[0]
                    #kuva = Image.open(kuva)
                    print(str(kalastaja)+" "+str(paikka)+" "+str(viehe)+" "+str(sp))
                    print(str(paino),str(pituus))
                    print(str(type(paino)),str(type(pituus)))                    
                    Saalis.objects.create(
                            saaja = kalastaja,
                            paikka = paikka,
                            paino = paino,
                            vapautettu = vapaus,
                            pituus = pituus,
                            viehe = viehe,
                            email = None,
                            saantipaiva = aika,
                            kuva = kuva,
                            public = True,
                            sukupuoli = sp)
                else:
                    if test == True:
                        return HttpResponse("Fail")
                    Saalis.objects.create(
                            saaja = kalastaja,
                            paikka = paikka,
                            paino = paino,
                            vapautettu = vapaus,
                            pituus = pituus,
                            viehe = viehe,
                            email = None,
                            saantipaiva = aika,
                            public = True,
                            sukupuoli = sp)
                """
                except Exception as e:
                    temp = str(type(kalastaja)) + " " + str(type(paikka)) + " " + str(type(paino)) + " " + str(type(pituus)) + " " + str(type(viehe)) + " " + str(type(kuva)) + "\n"
                    temp = temp + str(kalastaja) + " " + str(paikka) + " " + str(paino) + " " + str(pituus) + " " + str(viehe) + " " + str(aika) + " " + str(sp) + " " +  str(kuva) + " " + str(type(kuva)) + "\n" + repr(e) 
                    print(temp)
                    return HttpResponse(temp)
                """
                
        
        print("__________")
        for i in skip_paino:
            print(i,type(i))

        print("__________")
        for i in skip_pituus:
            print(i,type(i))

        print("__________")
        for i in sorted(null):
            print(i,null[i])

        print("__________")
        for puoli in sukupuolet:
            print(puoli)

        print("__________")
        for kala in wrong_kalat:
            print(kala)

        print("__________")
        for image in skip_image:
            print(image)

        print("__________")
        print("SKIP COUNT: " + str(skip_count))
        
        wb2.save(os.path.join(PROJECT_ROOT, 'faulty.xlsx'))
        #Saalis.objects.bulk_create(objects)

        return HttpResponse("Thank you for parseing with us")

def index(request):

    

    data = request.GET.copy()
    now = timezone.now()
    saaliit = Saalis.objects.filter(public=True)


    year = data.get('vuosi')

    if not data.get('vuosi') and data.get('Kalastajat') == None:
        year = int(now.year)
    elif not data.get('vuosi'):
        year = None
    elif data.get('vuosi') == '-':
        year = None
    else:
        year = int(data.get('vuosi'))



    #year = data.get('vuosi')
    date_from = data.get('date_from')
    date_to = data.get('date_to')
    kalastajat = data.get('Kalastajat')
    saantipaikka = data.get('Saantipaikka')
    viehe = data.get('Viehe')
    paino_from = data.get('Paino_from')
    paino_to = data.get('Paino_to')
    pituus_from = data.get('Pituus_from')
    pituus_to = data.get('Pituus_to')
    sort = data.get('sort')
    reverse = data.get('reverse')
    if not sort:
        sort = 'aika'
    p = data.get('p')
    

    
    

    years = [d.year for d in saaliit.dates('saantipaiva','year')]
    if now.year not in years:
        years.append(now.year)
    
   

    
    

    if p == "" or p == None:
        p = 100

    p = int(p)

    
    for i in range(0, len(years)):
        years[i] = int(years[i])
    years.sort(reverse = True)
    

    if year:
        saaliit = saaliit.filter(saantipaiva__year=year)
    if date_from:
        saaliit = saaliit.filter(saantipaiva__gte=date_from)
    if date_to:
        saaliit = saaliit.filter(saantipaiva__lte=date_to)
    if kalastajat:
        saaliit = saaliit.filter(saaja__contains=kalastajat)
    if saantipaikka:
        saaliit = saaliit.filter(paikka__contains=saantipaikka)
    if viehe:
        saaliit = saaliit.filter(viehe__contains=viehe)
    if paino_from:
        saaliit = saaliit.filter(paino__gte=paino_from)
    if paino_to:
        saaliit = saaliit.filter(paino__lte=paino_to)
    if pituus_from:
        saaliit = saaliit.filter(pituus__gte=pituus_from)
    if pituus_to:
        saaliit = saaliit.filter(pituus__lte=pituus_to)
    
    if sort == 'paino':
        if reverse:
            saaliit = saaliit.order_by('paino','id')
        else:
            saaliit = saaliit.order_by('-paino','-id')
    elif sort == 'pituus':
        if reverse:
            saaliit = saaliit.order_by('pituus','id')
        else:
            saaliit = saaliit.order_by('-pituus','-id')
    else:
        if reverse:
            saaliit = saaliit.order_by('saantipaiva','id')
        else:
            saaliit = saaliit.order_by('-saantipaiva','-id')

    
    pagination = Paginator(saaliit,p)
    p1 = pagination.page(1)

    #saaliit = Saalis.objects.filter(saantipaiva__year=now.year).order_by('saantipaiva')[:5]
    if reverse:
        reverse = True
    else:
        reverse = False

    template_name = 'saalistilastot.html'
    context = {
    'saaliit': p1.object_list,
    'paginatior': pagination,
    'years': years,
    'year': year,
    'date_from': data.get('date_from'),
    'date_to': data.get('date_to'),
    'kalastajat': data.get('Kalastajat'),
    'saantipaikka': data.get('Saantipaikka'),
    'viehe': data.get('Viehe'),
    'paino_from': data.get('Paino_from'),
    'paino_to': data.get('Paino_to'),
    'pituus_from': data.get('Pituus_from'),
    'pituus_to': data.get('Pituus_to'),
    'p': p,
    'sort': sort,
    'reverse': reverse,
    }
    return render(request,template_name,context)
    

def post(request):
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        form = SaalisForm(request.POST, request.FILES or None)
        if form.is_valid():
            instance = form.save(commit = False) 
            instance.save()
            return HttpResponse("Thanks")

        else:
            return HttpResponse("Invalid form")

    # if a GET (or any other method) we'll create a blank form
    else:
        form = SaalisForm()
    return render(request, 'form.html', {'form': form})

def biggest(request):
    saaliit = Saalis.objects.order_by('-paino')[:1].first()
    template = loader.get_template('suurin.html')
    context = {
        'suurin': saaliit,
    }
    return HttpResponse(template.render(context, request))


def table(request):
    saaliit = Saalis.objects.order_by('saantipaiva')[:5]
    template = loader.get_template('result.html')
    context = {
        'saaliit': saaliit,
    }
    return HttpResponse(template.render(context, request))

# Create your views here.
